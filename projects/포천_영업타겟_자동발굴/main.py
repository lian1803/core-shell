"""
포천 지역 영업 타겟 자동 발굴 도구

사용법: python main.py
결과:  output/포천_영업타겟_YYYYMMDD_HHMMSS.xlsx
"""

import asyncio
import re
import sys
import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set

# Windows 콘솔 UTF-8 출력 설정
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ─────────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────────
REGION = "포천"
TARGET_COUNT = 100

NAVER_CLIENT_ID = "o0776HJDmQVO6J9Lez1m"
NAVER_CLIENT_SECRET = "ZXG0lPbgH9"

MOBILE_UA = (
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
)

CATEGORIES = [
    "음식점", "카페", "미용실", "네일샵", "병원", "학원",
    "마트", "헬스장", "세탁소", "꽃집", "약국", "치과",
    "피부과", "노래방", "베이커리", "부동산", "숙박",
    "자동차", "인테리어", "옷가게",
]


# ─────────────────────────────────────────────────
# 1단계: 네이버 Local API로 업체 목록 수집
# ─────────────────────────────────────────────────
async def fetch_by_category(category: str) -> List[Dict]:
    """Naver Local Search API — 포천 특정 카테고리 업체 반환"""
    headers = {
        "X-Naver-Client-Id": NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }
    params = {"query": f"{REGION} {category}", "display": 5}
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.get(
                "https://openapi.naver.com/v1/search/local.json",
                headers=headers, params=params, timeout=10.0
            )
            resp.raise_for_status()
            items = resp.json().get("items", [])

        results = []
        for item in items:
            name = re.sub(r"<[^>]+>", "", item.get("title", "")).strip()
            tel = item.get("telephone", "")
            results.append({
                "name": name,
                "category": item.get("category", category),
                "address": item.get("roadAddress") or item.get("address", ""),
                "telephone": tel,
                "is_010": tel.startswith("010"),
                "detail": {},
                "daangn_url": "",
                "score": 50,
            })
        return results
    except Exception as e:
        print(f"  [API 오류] {category}: {e}")
        return []


# ─────────────────────────────────────────────────
# 2단계: Playwright로 네이버 플레이스 상세 수집
# ─────────────────────────────────────────────────
def _extract_place_id(text: str) -> Optional[str]:
    m = re.search(r'/place/(\d{6,})', text)
    return m.group(1) if m else None


def _parse_num(s: str) -> int:
    s = s.strip().replace(",", "")
    try:
        if "만" in s:
            return int(float(s.replace("만", "")) * 10000)
        if "천" in s:
            return int(float(s.replace("천", "")) * 1000)
        return int(float(s))
    except:
        return 0


def _extract_counts(text: str) -> Dict:
    result = {"photo_count": 0, "visitor_review": 0, "receipt_review": 0, "blog_review": 0}
    for key, pattern in [
        ("photo_count",    r'사진\s*([\d.,]+[만천]?)\s*장'),
        ("visitor_review", r'방문자\s*리뷰\s*([\d.,]+[만천]?)'),
        ("receipt_review", r'영수증\s*리뷰\s*([\d.,]+[만천]?)'),
        ("blog_review",    r'블로그\s*리뷰\s*([\d.,]+[만천]?)'),
    ]:
        m = re.search(pattern, text)
        if m:
            result[key] = _parse_num(m.group(1))
    return result


async def get_place_detail(search_page, photo_page, name: str, address: str) -> Dict:
    """
    1단계: 모바일 검색 → place_id + 리뷰/영업시간 등
    2단계: /photo 페이지 → 정확한 사진 수
    """
    detail = {
        "place_id": None, "naver_url": "",
        "photo_count": 0, "visitor_review": 0, "receipt_review": 0, "blog_review": 0,
        "has_hours": False, "has_menu": False, "has_navertalk": False, "has_instagram": False,
    }
    try:
        query = f"{name} {address[:10]}"
        encoded = urllib.parse.quote(query)
        await search_page.goto(
            f"https://m.search.naver.com/search.naver?query={encoded}&where=m_local",
            timeout=30000
        )
        await search_page.wait_for_load_state("networkidle", timeout=20000)
        await search_page.wait_for_timeout(1500)

        text = await search_page.inner_text("body")
        html = await search_page.content()

        pid = _extract_place_id(html)
        if pid:
            detail["place_id"] = pid
            detail["naver_url"] = f"https://map.naver.com/p/entry/place/{pid}"

        detail.update(_extract_counts(text))
        detail["has_hours"]     = bool(re.search(r'영업시간|운영시간', text))
        detail["has_menu"]      = bool(re.search(r'메뉴\s*\d+|메뉴판', text))
        detail["has_navertalk"] = "톡톡" in text or "talk.naver" in html
        detail["has_instagram"] = "instagram" in html.lower() or "인스타" in text

        # 2단계: place_id 있으면 /photo 페이지에서 사진 수 추출
        if pid:
            photo_count = await _fetch_photo_count(photo_page, pid)
            if photo_count > 0:
                detail["photo_count"] = photo_count

    except Exception as e:
        print(f"    ⚠ 상세 오류 ({name}): {e}")
    return detail


async def _fetch_photo_count(page, place_id: str) -> int:
    """
    m.place.naver.com/place/{id}/photo 에서 사진 수 추출
    페이지 구조: "이미지 갯수\n999+\n..." 또는 "이미지 갯수\n1,234\n..."
    """
    try:
        await page.goto(
            f"https://m.place.naver.com/place/{place_id}/photo",
            timeout=20000
        )
        await page.wait_for_load_state("domcontentloaded", timeout=15000)
        await page.wait_for_timeout(1000)
        text = await page.inner_text("body")

        # "이미지 갯수" 다음에 오는 숫자 (예: "999+", "1,234")
        m = re.search(r'이미지\s*갯수\s*[\r\n\s]*([\d,]+)\+?', text)
        if m:
            return _parse_num(m.group(1))

        # 독립형 "999+" 패턴
        m = re.search(r'^([\d,]+)\+?\s*$', text, re.MULTILINE)
        if m:
            count = _parse_num(m.group(1))
            if count > 0:
                return count

        # fallback: 큰 숫자 하나 (사진이 많은 경우 첫 큰 수)
        nums = re.findall(r'([\d,]{3,})', text)
        for n in nums:
            count = _parse_num(n)
            if 10 <= count <= 9999:
                return count
    except:
        pass
    return 0


# ─────────────────────────────────────────────────
# 3단계: 당근마켓 검색 링크 생성 (크롤링 없이)
# ─────────────────────────────────────────────────
def make_daangn_search_url(name: str) -> str:
    """업체명으로 당근마켓 검색 URL 생성 (클릭 후 수동 확인용)"""
    encoded = urllib.parse.quote(name)
    return f"https://www.daangn.com/kr/search/?q={encoded}"


# ─────────────────────────────────────────────────
# 4단계: 우선순위 점수 계산
# ─────────────────────────────────────────────────
def calc_score(detail: Dict, is_010: bool) -> int:
    """낮을수록 = 더 취약 = 더 좋은 영업 타겟"""
    score = 100
    photos  = detail.get("photo_count", 0)
    reviews = detail.get("visitor_review", 0) + detail.get("receipt_review", 0)
    blog    = detail.get("blog_review", 0)

    if photos == 0:        score -= 30
    elif photos < 5:       score -= 15
    elif photos < 20:      score -= 5

    if reviews == 0:       score -= 25
    elif reviews < 10:     score -= 12
    elif reviews < 30:     score -= 5

    if blog == 0:          score -= 15
    elif blog < 5:         score -= 7

    if not detail.get("has_hours"):  score -= 10
    if not detail.get("has_menu"):   score -= 10
    if is_010:                       score -= 5   # 010 개인번호 = 추가 감점

    return max(0, score)


def priority_label(score: int) -> str:
    if score <= 20:  return "🔥 최우선"
    if score <= 40:  return "⚡ 우선"
    if score <= 60:  return "✅ 보통"
    return "⬜ 낮음"


# ─────────────────────────────────────────────────
# 5단계: 엑셀 저장
# ─────────────────────────────────────────────────
def save_excel(businesses: List[Dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "영업 타겟"

    headers = [
        "순위", "우선순위", "업체명", "업종", "주소",
        "사진수", "방문리뷰", "영수증리뷰", "블로그리뷰",
        "영업시간", "메뉴정보", "네이버톡톡", "인스타",
        "당근검색링크(클릭확인)", "전화번호", "010여부",
        "점수", "네이버플레이스링크",
    ]

    header_fill = PatternFill("solid", fgColor="1a365d")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    ws.row_dimensions[1].height = 28
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    color_map = {
        "🔥 최우선": "ff6b6b",
        "⚡ 우선":   "ffd93d",
        "✅ 보통":   "c6efce",
        "⬜ 낮음":   "f5f5f5",
    }

    for idx, biz in enumerate(businesses, 1):
        d     = biz.get("detail", {})
        label = priority_label(biz.get("score", 50))
        fill  = PatternFill("solid", fgColor=color_map.get(label, "ffffff"))

        row_data = [
            idx, label,
            biz.get("name", ""), biz.get("category", ""), biz.get("address", ""),
            d.get("photo_count", 0), d.get("visitor_review", 0),
            d.get("receipt_review", 0), d.get("blog_review", 0),
            "O" if d.get("has_hours") else "X",
            "O" if d.get("has_menu")  else "X",
            "O" if d.get("has_navertalk") else "X",
            "O" if d.get("has_instagram") else "X",
            biz.get("daangn_url", ""),
            biz.get("telephone", ""),
            "O" if biz.get("is_010") else "X",
            biz.get("score", 50),
            d.get("naver_url", ""),
        ]

        row_num = idx + 1
        ws.row_dimensions[row_num].height = 22
        left_cols = {3, 4, 5, 14, 18}
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = fill
            cell.alignment = left if col in left_cols else center

    col_widths = [5, 12, 18, 12, 28, 6, 8, 8, 8, 8, 8, 8, 6, 42, 14, 6, 6, 48]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    Path(path).parent.mkdir(exist_ok=True)
    wb.save(path)


# ─────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────
async def main():
    print("=" * 58)
    print("  포천 지역 영업 타겟 자동 발굴 도구")
    print(f"  목표: {TARGET_COUNT}개 업체 | 전 업종 | 엑셀 출력")
    print("=" * 58)

    # ── 1단계: 업체 목록 수집 ──────────────────────
    print("\n📍 [1/4] 네이버 API로 업체 목록 수집 중...\n")
    businesses: List[Dict] = []
    seen: Set[str] = set()

    for cat in CATEGORIES:
        if len(businesses) >= TARGET_COUNT:
            break
        items = await fetch_by_category(cat)
        added = 0
        for item in items:
            if item["name"] not in seen:
                seen.add(item["name"])
                businesses.append(item)
                added += 1
        print(f"  {REGION} {cat:<8}: +{added}개  (누적 {len(businesses)}개)")
        await asyncio.sleep(0.3)

    businesses = businesses[:TARGET_COUNT]
    print(f"\n  ✅ {len(businesses)}개 업체 확보")

    # ── 2·3단계: 상세 분석 + 당근 확인 ───────────
    print("\n📊 [2/4] 네이버 플레이스 품질 분석 중...\n")
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)

        mobile_ctx  = await browser.new_context(
            user_agent=MOBILE_UA, viewport={"width": 390, "height": 844}, locale="ko-KR"
        )
        search_page = await mobile_ctx.new_page()
        photo_page  = await mobile_ctx.new_page()   # 사진 수 전용 탭

        for i, biz in enumerate(businesses, 1):
            label_now = f"[{i:3d}/{len(businesses)}] {biz['name'][:16]:<16}"
            print(f"  {label_now}", end=" ", flush=True)

            detail = await get_place_detail(search_page, photo_page, biz["name"], biz["address"])
            biz["detail"] = detail
            biz["score"]  = calc_score(detail, biz.get("is_010", False))
            # 당근마켓 검색 URL (클릭해서 직접 확인)
            biz["daangn_url"] = make_daangn_search_url(biz["name"])
            print(
                f"사진:{detail['photo_count']:3d}  리뷰:{detail['visitor_review']:4d}"
                f"  → {priority_label(biz['score'])}"
            )
            await asyncio.sleep(0.8)

        await mobile_ctx.close()
        await browser.close()

    # ── 4단계: 정렬 & 저장 ─────────────────────────
    businesses.sort(key=lambda x: x.get("score", 50))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = f"output/포천_영업타겟_{timestamp}.xlsx"

    print("\n💾 [3/4] 엑셀 저장 중...")
    save_excel(businesses, out_path)

    hot  = sum(1 for b in businesses if b.get("score", 50) <= 20)
    warm = sum(1 for b in businesses if 20 < b.get("score", 50) <= 40)

    print(f"\n{'='*58}")
    print(f"  완료!  파일: {out_path}")
    print(f"  최우선: {hot}개   우선: {warm}개")
    print(f"  당근마켓: 엑셀 링크 클릭해서 직접 확인")
    print(f"{'='*58}")


if __name__ == "__main__":
    asyncio.run(main())
