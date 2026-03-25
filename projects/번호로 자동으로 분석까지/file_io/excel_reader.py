"""
io/excel_reader.py — 엑셀에서 업체 정보 읽기 (단건 + 배치)
"""
import sys
import os
import re
import urllib.parse
from typing import Optional, Dict, List

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import log


def _extract_naver_url(row_dict: Dict) -> str:
    """row dict에서 네이버 플레이스 URL 추출"""
    # 지정 컬럼명 우선
    for col_name in ["네이버플레이스", "플레이스주소", "주소", "url", "URL"]:
        val = str(row_dict.get(col_name, "") or "")
        if val.startswith("http") and "naver" in val and "place" in val:
            return val

    # 전체 값에서 탐색
    for val in row_dict.values():
        val_str = str(val or "")
        if val_str.startswith("http") and "naver" in val_str and "place" in val_str:
            return val_str

    return ""


def _infer_main_keyword(row_dict: Dict, place_url: str, category: str) -> str:
    """메인 키워드 추론 (지역 + 업종)"""
    # 엑셀에 메인키워드 컬럼이 있으면 사용
    for col in ["메인키워드", "키워드", "검색키워드"]:
        val = str(row_dict.get(col, "") or "").strip()
        if val:
            return val

    # URL에서 지역명 추출
    city = ""
    decoded_url = urllib.parse.unquote(place_url)
    city_match = re.search(r"search/([가-힣]+)\s+[가-힣]+/place", decoded_url)
    if city_match:
        city = city_match.group(1)

    if not city:
        st_match = re.search(r"searchText=([^&]+)", place_url)
        if st_match:
            st = urllib.parse.unquote(st_match.group(1))
            parts = st.split()
            if len(parts) >= 1:
                city = parts[0]

    return f"{city} {category}".strip() if city and category else category or ""


def read_first_url_row(path: str) -> Optional[Dict]:
    """엑셀에서 첫 번째 네이버 플레이스 URL이 있는 행 반환"""
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [str(c.value) for c in ws[1]]
        log("Excel", f"컬럼: {headers}")

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            url = _extract_naver_url(row_dict)
            if url:
                category = str(row_dict.get("업종", "") or "")
                row_dict["_place_url"] = url
                row_dict["_main_keyword"] = _infer_main_keyword(row_dict, url, category)
                log("Excel", f"발견: {row_dict.get('업체명')} | URL: {url[:60]}...")
                return row_dict

        log("Excel", "URL 있는 행 없음")
        return None
    except Exception as e:
        log("Excel", f"읽기 오류: {e}")
        return None


def read_all_businesses(path: str) -> List[Dict]:
    """엑셀 전체 행에서 네이버 URL이 있는 모든 업체 반환"""
    results = []
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [str(c.value) for c in ws[1]]

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = dict(zip(headers, row))
            url = _extract_naver_url(row_dict)
            if url:
                category = str(row_dict.get("업종", "") or "")
                row_dict["_place_url"] = url
                row_dict["_main_keyword"] = _infer_main_keyword(row_dict, url, category)
                row_dict["_row_num"] = i
                results.append(row_dict)

        log("Excel", f"총 {len(results)}개 업체 발견")
    except Exception as e:
        log("Excel", f"읽기 오류: {e}")

    return results


def extract_place_id(url: str) -> Optional[str]:
    """URL에서 place_id 추출"""
    match = re.search(r"place/(\d+)", url)
    return match.group(1) if match else None
