# -*- coding: utf-8 -*-
"""엑셀에서 주문 읽기"""

from dataclasses import dataclass
from pathlib import Path
from openpyxl import load_workbook


@dataclass
class Order:
    row_num: int
    market: str          # 주문마켓 (06.쿠팡, 03.11번가 등)
    product_name: str    # 품목
    option: str          # 옵션
    quantity: int        # 수량
    buyer: str           # 구매자
    seller_code: str     # 업체코드 (불사자 검색키)
    product_code: str    # 업체코드에서 콜론 앞 (상품 고유코드)
    status: str          # 정산금 옆 상태


def load_orders(excel_path: str | Path, sheet_name: str = "Sheet1") -> list[Order]:
    """엑셀에서 주문 목록 읽기"""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    orders = []
    header = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # 헤더 찾기
        if header is None:
            if "업체코드" in row or "품목" in row:
                header = list(row)
            continue

        # 빈 행 스킵
        if not any(row):
            continue

        def get(col_name):
            try:
                idx = header.index(col_name)
                return str(row[idx]) if row[idx] is not None else ""
            except ValueError:
                return ""

        seller_code = get("업체코드")
        if not seller_code:
            continue

        # 업체코드에서 콜론 앞 = 상품코드
        product_code = seller_code.split(":")[0] if ":" in seller_code else seller_code

        orders.append(Order(
            row_num=i + 1,
            market=get("주문마켓"),
            product_name=get("품목"),
            option=get("옵션"),
            quantity=int(get("수량") or 1),
            buyer=get("구매자"),
            seller_code=seller_code,
            product_code=product_code,
            status=get("정산금"),
        ))

    wb.close()
    return orders
