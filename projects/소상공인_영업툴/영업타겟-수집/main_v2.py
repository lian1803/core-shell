"""
포천 지역 영업 타겟 자동 발굴 도구 v2
- 네이버 플레이스 전화번호 + 품질 분석
- 인스타그램 계정 URL 자동 수집
- 당근마켓 비즈프로필 URL 수집 (네이버 검색 경유)
- 대형 프랜차이즈 자동 제외

사용법: python main_v2.py
결과:  output/포천_영업타겟_v2_YYYYMMDD_HHMMSS.xlsx
"""

import asyncio
import re
import sys
import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright

# ─────────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────────
REGION = "포천"
TARGET_COUNT = 100

NAVER_CLIENT_ID = "o0776HJDmQVO6J9Lez1m"
NAVER_CLIENT_SECRET = "ZXG0lPbgH9"

MOBILE_UA = (
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
)
DESKTOP_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

CATEGORIES = [
    "음식점", "카페", "미용실", "네일샵", "병원", "학원",
    "마트", "헬스장", "세탁소", "꽃집", "약국", "치과",
    "피부과", "노래방", "베이커리", "부동산", "숙박",
    "자동차", "인테리어", "옷가게",
]

# 대형 프랜차이즈 / 공공기관 — 영업 타겟 제외
FRANCHISE_EXCLUDE = {
    "스타벅스", "이마트", "홈플러스", "롯데마트", "코스트코", "다이소",
    "맥도날드", "버거킹", "롯데리아", "KFC", "서브웨이", "맘스터치",
    "GS25", "CU", "세븐일레븐", "미니스톱", "이마트24",
    "올리브영", "다이소", "무신사", "탑텐", "유니클로", "자라", "H&M",
    "크린토피아",  # 프랜차이즈 세탁소
    "CHA의과학대학교", "경기도의료원", "국군", "보훈",
    "한화리조트", "대명리조트",
    "국군복지단", "하나로마트",
}

IG_PATTERN = re.compile(
    r'https?://(?:www\.)?instagram\.com/([A-Za-z0-9._]{2,30})(?:/[^"\s]*)?'
)
PHONE_PATTERN = re.compile(r'0\d{1,2}-\d{3,4}-\d{4}')


# ─────────────────────────────────────────────────
# 1단계: 네이버 Local API로 업체 목록 수집
# ─────────────────────────────────────────────────
async def fetch_by_category(category: str) -> List[Dict]:
    headers = {
        "X-Naver-Client-Id": NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }
    params = {"query": f"{REGION} {category}", "display": 5}
    try:
        async with httpx.AsyncClient() as client:
            resp = await client.get(
                "https://openapi.naver.com/v1/search/local.json",
                headers=headers, params=params, timeout=10.0
            )
            resp.raise_for_status()
            items = resp.json().get("items", [])

        results = []
        for item in items:
            name = re.sub(r"<[^>]+>", "", item.get("title", "")).strip()
            tel = item.get("telephone", "")
            results.append({
                "name": name,
                "category": item.get("category", category),
                "address": item.get("roadAddress") or item.get("address", ""),
                "telephone": tel,
                "is_010": tel.startswith("010"),
                "detail": {},
                "instagram_url": "",
                "daangn_url": "",
                "daangn_phone": "",
                "score": 50,
            })
        return results
    except Exception as e:
        print(f"  [API 오류] {category}: {e}")
        return []


# ─────────────────────────────────────────────────
# 2단계: 네이버 플레이스 상세 수집
# ─────────────────────────────────────────────────
def _extract_place_id(text: str) -> Optional[str]:
    m = re.search(r'/place/(\d{6,})', text)
    return m.group(1) if m else None


def _parse_num(s: str) -> int:
    s = s.strip().replace(",", "")
    try:
        if "만" in s:
            return int(float(s.replace("만", "")) * 10000)
        if "천" in s:
            return int(float(s.replace("천", "")) * 1000)
        return int(float(s))
    except:
        return 0


def _extract_counts(text: str) -> Dict:
    result = {"photo_count": 0, "visitor_review": 0, "receipt_review": 0, "blog_review": 0}
    for key, pattern in [
        ("photo_count",    r'사진\s*([\d.,]+[만천]?)\s*장'),
        ("visitor_review", r'방문자\s*리뷰\s*([\d.,]+[만천]?)'),
        ("receipt_review", r'영수증\s*리뷰\s*([\d.,]+[만천]?)'),
        ("blog_review",    r'블로그\s*리뷰\s*([\d.,]+[만천]?)'),
    ]:
        m = re.search(pattern, text)
        if m:
            result[key] = _parse_num(m.group(1))
    return result


def _extract_instagram_from_html(html: str) -> str:
    """HTML에서 instagram.com URL 추출. 공식 계정만 (explore, p/, reel/ 등 제외)"""
    EXCLUDE = {"explore", "p", "reel", "reels", "stories", "accounts", "tv",
               "static.cdninstagram.com", "about", "help", "legal"}
    for m in IG_PATTERN.finditer(html):
        username = m.group(1).rstrip("/")
        if username.lower() not in EXCLUDE and not username.startswith("_"):
            return f"https://www.instagram.com/{username}/"
    return ""


async def get_place_detail(search_page, photo_page, name: str, address: str) -> Dict:
    detail = {
        "place_id": None, "naver_url": "",
        "telephone": "",
        "photo_count": 0, "visitor_review": 0, "receipt_review": 0, "blog_review": 0,
        "has_hours": False, "has_menu": False, "has_navertalk": False,
        "instagram_url": "",
    }
    try:
        query = f"{name} {address[:10]}"
        encoded = urllib.parse.quote(query)
        await search_page.goto(
            f"https://m.search.naver.com/search.naver?query={encoded}&where=m_local",
            timeout=30000
        )
        await search_page.wait_for_load_state("networkidle", timeout=20000)
        await search_page.wait_for_timeout(1500)

        text = await search_page.inner_text("body")
        html = await search_page.content()

        pid = _extract_place_id(html)

        # fallback: 업체명만으로 재검색
        if not pid:
            await search_page.goto(
                f"https://m.search.naver.com/search.naver?query={urllib.parse.quote(name)}&where=m_local",
                timeout=20000
            )
            await search_page.wait_for_load_state("networkidle", timeout=15000)
            await search_page.wait_for_timeout(1000)
            html = await search_page.content()
            text = await search_page.inner_text("body")
            pid = _extract_place_id(html)

        if pid:
            detail["place_id"] = pid
            detail["naver_url"] = f"https://map.naver.com/p/entry/place/{pid}"

        detail.update(_extract_counts(text))
        detail["has_hours"]     = bool(re.search(r'영업시간|운영시간', text))
        detail["has_menu"]      = bool(re.search(r'메뉴\s*\d+|메뉴판', text))
        detail["has_navertalk"] = "톡톡" in text or "talk.naver" in html

        # 전화번호 추출 (API가 안 줄 때 페이지 텍스트에서)
        phone_match = PHONE_PATTERN.search(text)
        if phone_match:
            detail["telephone"] = phone_match.group(0)

        # 인스타그램 URL 추출 (플레이스 페이지에서 직접)
        ig = _extract_instagram_from_html(html)
        if ig:
            detail["instagram_url"] = ig

        # 사진 수: /photo 페이지에서 정확히
        if pid:
            photo_count = await _fetch_photo_count(photo_page, pid)
            if photo_count > 0:
                detail["photo_count"] = photo_count

    except Exception as e:
        print(f"    ⚠ 플레이스 오류 ({name}): {e}")
    return detail


async def _fetch_photo_count(page, place_id: str) -> int:
    try:
        await page.goto(
            f"https://m.place.naver.com/place/{place_id}/photo",
            timeout=20000
        )
        await page.wait_for_load_state("domcontentloaded", timeout=15000)
        await page.wait_for_timeout(1000)
        text = await page.inner_text("body")

        m = re.search(r'이미지\s*갯수\s*[\r\n\s]*([\d,]+)\+?', text)
        if m:
            return _parse_num(m.group(1))
        m = re.search(r'^([\d,]+)\+?\s*$', text, re.MULTILINE)
        if m:
            count = _parse_num(m.group(1))
            if count > 0:
                return count
        nums = re.findall(r'([\d,]{3,})', text)
        for n in nums:
            count = _parse_num(n)
            if 10 <= count <= 9999:
                return count
    except:
        pass
    return 0


# ─────────────────────────────────────────────────
# 3단계: 인스타그램 계정 찾기 (네이버 검색 fallback)
# ─────────────────────────────────────────────────
async def find_instagram_via_naver(page, name: str, address: str) -> str:
    """네이버에서 '{업체명} {지역} 인스타그램' 검색 → instagram.com URL 추출"""
    try:
        query = f"{name} {REGION} 인스타그램"
        encoded = urllib.parse.quote(query)
        await page.goto(
            f"https://search.naver.com/search.naver?query={encoded}&where=web",
            timeout=20000
        )
        await page.wait_for_load_state("domcontentloaded", timeout=15000)
        await page.wait_for_timeout(800)

        html = await page.content()
        ig = _extract_instagram_from_html(html)
        return ig
    except:
        return ""


# ─────────────────────────────────────────────────
# 4단계: 당근마켓 비즈프로필 수집 (네이버 검색 경유)
# 당근 직접 크롤링은 봇 차단이 강해서,
# 네이버에서 "{업체명} 당근마켓" 검색 → daangn.com URL 추출
# ─────────────────────────────────────────────────
DAANGN_URL_PATTERN = re.compile(
    r'https?://(?:www\.)?daangn\.com/(?:kr/biz/|articles/|fleamarket/)[^\s"\'<>]+'
)

async def find_daangn_biz(page, name: str) -> Dict:
    """네이버 검색으로 당근마켓 비즈프로필 URL 찾기"""
    result = {"daangn_url": "", "daangn_phone": ""}
    fallback = f"https://www.daangn.com/kr/search/?q={urllib.parse.quote(name)}"
    try:
        query = f'"{name}" 당근마켓'
        encoded = urllib.parse.quote(query)
        await page.goto(
            f"https://search.naver.com/search.naver?query={encoded}&where=web",
            timeout=20000
        )
        await page.wait_for_load_state("domcontentloaded", timeout=15000)
        await page.wait_for_timeout(600)

        html = await page.content()

        # 당근마켓 비즈프로필 or 게시글 URL 추출
        matches = DAANGN_URL_PATTERN.findall(html)
        # 비즈프로필 (/kr/biz/) 우선
        biz_urls = [u for u in matches if "/kr/biz/" in u]
        other_urls = [u for u in matches if "/kr/biz/" not in u]

        if biz_urls:
            result["daangn_url"] = biz_urls[0].split('"')[0].split("'")[0]
        elif other_urls:
            result["daangn_url"] = other_urls[0].split('"')[0].split("'")[0]
        else:
            result["daangn_url"] = fallback

    except:
        result["daangn_url"] = fallback
    return result


# ─────────────────────────────────────────────────
# 5단계: 우선순위 점수 계산
# ─────────────────────────────────────────────────
def is_franchise(name: str) -> bool:
    """대형 프랜차이즈/공공기관 여부"""
    for keyword in FRANCHISE_EXCLUDE:
        if keyword in name:
            return True
    return False


def calc_score(detail: Dict, is_010: bool, franchise: bool = False) -> int:
    """낮을수록 = 더 취약 = 더 좋은 영업 타겟. 프랜차이즈는 999(제외)"""
    if franchise:
        return 999
    score = 100
    photos  = detail.get("photo_count", 0)
    reviews = detail.get("visitor_review", 0) + detail.get("receipt_review", 0)
    blog    = detail.get("blog_review", 0)

    if photos == 0:      score -= 30
    elif photos < 5:     score -= 15
    elif photos < 20:    score -= 5

    if reviews == 0:     score -= 25
    elif reviews < 10:   score -= 12
    elif reviews < 30:   score -= 5

    if blog == 0:        score -= 15
    elif blog < 5:       score -= 7

    if not detail.get("has_hours"):   score -= 10
    if not detail.get("has_menu"):    score -= 10
    if is_010:                        score -= 5

    return max(0, score)


def priority_label(score: int) -> str:
    if score == 999: return "❌ 제외(프랜차이즈)"
    if score <= 20:  return "🔥 최우선"
    if score <= 40:  return "⚡ 우선"
    if score <= 60:  return "✅ 보통"
    return "⬜ 낮음"


# ─────────────────────────────────────────────────
# 6단계: 엑셀 저장
# ─────────────────────────────────────────────────
def save_excel(businesses: List[Dict], path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "영업 타겟"

    headers = [
        "순위", "우선순위", "업체명", "업종", "주소",
        "전화번호", "010여부",
        "사진수", "방문리뷰", "영수증리뷰", "블로그리뷰",
        "영업시간", "메뉴정보", "네이버톡톡",
        "인스타그램",
        "당근마켓",
        "점수", "네이버플레이스",
    ]

    header_fill = PatternFill("solid", fgColor="1a365d")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    ws.row_dimensions[1].height = 28
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    color_map = {
        "🔥 최우선":        "ff6b6b",
        "⚡ 우선":          "ffd93d",
        "✅ 보통":          "c6efce",
        "⬜ 낮음":          "f5f5f5",
        "❌ 제외(프랜차이즈)": "cccccc",
    }

    for idx, biz in enumerate(businesses, 1):
        d     = biz.get("detail", {})
        label = priority_label(biz.get("score", 50))
        fill  = PatternFill("solid", fgColor=color_map.get(label, "ffffff"))

        ig_url     = biz.get("instagram_url", "") or d.get("instagram_url", "")
        daangn_url = biz.get("daangn_url", "")

        row_data = [
            idx, label,
            biz.get("name", ""), biz.get("category", ""), biz.get("address", ""),
            biz.get("telephone", ""),
            "O" if biz.get("is_010") else "X",
            d.get("photo_count", 0), d.get("visitor_review", 0),
            d.get("receipt_review", 0), d.get("blog_review", 0),
            "O" if d.get("has_hours") else "X",
            "O" if d.get("has_menu")  else "X",
            "O" if d.get("has_navertalk") else "X",
            ig_url,
            daangn_url,
            biz.get("score", 50),
            d.get("naver_url", ""),
        ]

        row_num = idx + 1
        ws.row_dimensions[row_num].height = 22
        url_cols = {15, 16, 18}   # 인스타, 당근, 네이버플레이스
        left_cols = {3, 4, 5} | url_cols

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill = fill
            cell.alignment = left if col in left_cols else center
            # URL 컬럼에 하이퍼링크
            if col in url_cols and val and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(color="0563C1", underline="single")

    col_widths = [5, 20, 18, 12, 28, 14, 6, 6, 8, 8, 8, 8, 8, 8, 38, 42, 6, 48]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    Path(path).parent.mkdir(exist_ok=True)
    wb.save(path)


# ─────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────
async def main():
    print("=" * 62)
    print("  포천 지역 영업 타겟 자동 발굴 도구 v2")
    print(f"  목표: {TARGET_COUNT}개 업체 | 전화 + 인스타 + 당근 수집")
    print("=" * 62)

    # ── 1단계: 업체 목록 수집 ──────────────────────
    print("\n📍 [1/4] 네이버 API로 업체 목록 수집 중...\n")
    businesses: List[Dict] = []
    seen: Set[str] = set()

    for cat in CATEGORIES:
        if len(businesses) >= TARGET_COUNT:
            break
        items = await fetch_by_category(cat)
        added = 0
        for item in items:
            if item["name"] not in seen:
                seen.add(item["name"])
                businesses.append(item)
                added += 1
        print(f"  {REGION} {cat:<8}: +{added}개  (누적 {len(businesses)}개)")
        await asyncio.sleep(0.3)

    businesses = businesses[:TARGET_COUNT]
    print(f"\n  ✅ {len(businesses)}개 업체 확보")

    # ── 2·3·4단계: 플레이스 + 인스타 + 당근 수집 ──
    print("\n📊 [2/4] 네이버 플레이스 + 인스타 + 당근마켓 수집 중...\n")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)

        # 모바일 컨텍스트 (네이버 플레이스용)
        mobile_ctx  = await browser.new_context(
            user_agent=MOBILE_UA,
            viewport={"width": 390, "height": 844},
            locale="ko-KR"
        )
        search_page = await mobile_ctx.new_page()
        photo_page  = await mobile_ctx.new_page()

        # 데스크탑 컨텍스트 (네이버 검색 + 당근용)
        desktop_ctx   = await browser.new_context(
            user_agent=DESKTOP_UA,
            viewport={"width": 1280, "height": 800},
            locale="ko-KR"
        )
        naver_srch_page = await desktop_ctx.new_page()
        daangn_page     = await desktop_ctx.new_page()

        for i, biz in enumerate(businesses, 1):
            label_now = f"[{i:3d}/{len(businesses)}] {biz['name'][:14]:<14}"
            print(f"  {label_now}", end=" ", flush=True)

            # 네이버 플레이스 상세
            detail = await get_place_detail(search_page, photo_page, biz["name"], biz["address"])
            biz["detail"]    = detail
            biz["franchise"] = is_franchise(biz["name"])
            biz["score"]     = calc_score(detail, biz.get("is_010", False), biz["franchise"])

            # 인스타그램: 플레이스에서 못 찾으면 네이버 검색
            ig = detail.get("instagram_url", "")
            if not ig:
                ig = await find_instagram_via_naver(naver_srch_page, biz["name"], biz["address"])
            biz["instagram_url"] = ig

            # 당근마켓
            daangn = await find_daangn_biz(daangn_page, biz["name"])
            biz["daangn_url"]   = daangn["daangn_url"]
            biz["daangn_phone"] = daangn["daangn_phone"]

            # 출력 상태
            ig_mark     = "📸" if ig else "  "
            daangn_mark = "🥕" if daangn["daangn_phone"] else ("🥕?" if "/kr/biz/" in daangn["daangn_url"] else "  ")
            print(
                f"사진:{detail['photo_count']:3d}  리뷰:{detail['visitor_review']:4d}"
                f"  {ig_mark}{daangn_mark}  → {priority_label(biz['score'])}"
            )
            await asyncio.sleep(0.5)

        await mobile_ctx.close()
        await desktop_ctx.close()
        await browser.close()

    # ── 정렬 & 저장 ───────────────────────────────
    businesses.sort(key=lambda x: x.get("score", 50))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = f"output/포천_영업타겟_v2_{timestamp}.xlsx"

    print("\n💾 [3/4] 엑셀 저장 중...")
    save_excel(businesses, out_path)

    real = [b for b in businesses if b.get("score", 999) != 999]
    hot     = sum(1 for b in real if b.get("score", 50) <= 20)
    warm    = sum(1 for b in real if 20 < b.get("score", 50) <= 40)
    excluded = len(businesses) - len(real)
    has_ig  = sum(1 for b in real if b.get("instagram_url"))
    has_dg  = sum(1 for b in real if b.get("daangn_url") and "search" not in b.get("daangn_url", ""))

    print(f"\n{'='*62}")
    print(f"  완료!  파일: {out_path}")
    print(f"  유효 업체: {len(real)}개  (프랜차이즈 제외: {excluded}개)")
    print(f"  최우선: {hot}개   우선: {warm}개")
    print(f"  인스타 계정 발견: {has_ig}개 / {len(real)}개")
    print(f"  당근 비즈프로필 발견: {has_dg}개 / {len(real)}개")
    print(f"{'='*62}")


if __name__ == "__main__":
    asyncio.run(main())
